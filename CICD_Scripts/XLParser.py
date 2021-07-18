#-----------------------------------------------------------
#-----------------------------------------------------------
#-------------@Author: nishant.kumar_3@philips.com----------
#-----------------------------------------------------------
#-----------------------------------------------------------

import openpyxl
from openpyxl import load_workbook

class xl():
    def xl_Parse(self):
        try:
            wb = openpyxl.Workbook()
            wb = openpyxl.load_workbook(filename ='C:\\Users\\320035688\\Philips\\OneDrive - Philips\\Documents\\PropertiesDump.xlsx')
            sheet = wb[sheetname('sheet')]
            i = 2
            print(sheet['B'+str(i)].value)
            while (str(sheet['B'+str(i)].value) != "None"):
                #print(sheet['B'+str(i)].value)
                i = i+1
            count = i-2
            lis = [[] for x in range(count)]
            print (lis)
            for num in range(0,count):
                #print ("****",num)
                #This list needs to be appended for each column addition in XL sheet
                lis[num].append(sheet['B'+str(num+2)].value)
                lis[num].append(sheet['C'+str(num+2)].value)
                lis[num].append(sheet['D'+str(num+2)].value)
                lis[num].append(sheet['E'+str(num+2)].value)
                lis[num].append(sheet['F'+str(num+2)].value)
                lis[num].append(sheet['G'+str(num+2)].value)
                lis[num].append(sheet['H'+str(num+2)].value)
                lis[num].append(sheet['I'+str(num+2)].value)
                lis[num].append(sheet['J'+str(num+2)].value)
                lis[num].append(sheet['K'+str(num+2)].value)
            print(lis)
            #for each in range(count):
                #print(each)
                #print(lis[each])
                              
            return count, lis
        except exception as e:
            print (e)
'''
if __name__== "__main__":
    count, lis = xl_Parse()
    print("----------",count)
    print("---------->",lis)
    
'''

obj1 = xl()
lis = obj1.xl_Parse()
print(lis)